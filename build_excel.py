"""
build_excel.py
--------------
Assembles the final Excel deliverable: "Waste_Dump_Slope_Stability_Assessment.xlsx"

Sheets:
  1. Inputs & Baseline   - editable baseline parameters (blue) + live FS formula
  2. Sensitivity         - OAT sweep tables (live formulas) + tornado summary + charts
  3. Monte Carlo         - distribution assumptions, 2,000-row live sample, and the
                           full n=100,000 simulation summary (Python-computed, clearly labeled)
  4. Risk Optimization   - Pf vs slope angle design table + recommended angles + chart
  5. Notes & Methodology - assumptions, equation, limitations
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from core import BASELINE
from sensitivity import PARAM_RANGES, PARAM_LABELS, N_STEPS, tornado_data
from monte_carlo import default_distributions, run_simulation, summarize
from optimization import sweep_slope_angle, recommended_angles, TARGET_PF

FONT_NAME = "Calibri"
BLUE_INPUT = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

OUT_DIR = "../figures"  # figures were generated here by make_plots.py
WB_PATH = "../Waste_Dump_Slope_Stability_Assessment.xlsx"  # workbook lives at repo root


def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ----------------------------------------------------------------------------
# Sheet 1: Inputs & Baseline
# ----------------------------------------------------------------------------
def build_inputs_sheet(wb):
    ws = wb.active
    ws.title = "Inputs & Baseline"

    ws["A1"] = "Probabilistic Slope Stability Assessment - Mine Waste Rock Dump"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = "Model: Infinite-slope method with pore-pressure ratio r_u (Bishop & Morgenstern, 1960)"
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10)
    ws.merge_cells("A2:D2")

    ws["A4"] = "Baseline Input Parameters"
    ws["A4"].font = BOLD
    headers = ["Parameter", "Value", "Unit", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 1, 4)

    rows = [
        ("Cohesion, c'", BASELINE.cohesion_kPa, "kPa", "Effective cohesion of dump material"),
        ("Friction angle, phi'", BASELINE.friction_angle_deg, "deg", "Effective friction angle of dump material"),
        ("Unit weight, gamma", BASELINE.unit_weight_kNm3, "kN/m^3", "Bulk unit weight of waste rock"),
        ("Depth to slip surface, z", BASELINE.depth_m, "m", "Vertical depth to potential failure plane"),
        ("Slope angle, beta", BASELINE.slope_angle_deg, "deg", "Overall face angle of the dump"),
        ("Pore pressure ratio, r_u", BASELINE.ru, "-", "u / (gamma * z), Bishop & Morgenstern (1960)"),
    ]
    start_row = 6
    for i, (label, val, unit, note) in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=label).border = BORDER
        c = ws.cell(row=r, column=2, value=val)
        c.font = BLUE_INPUT
        c.fill = YELLOW_FILL
        c.border = BORDER
        ws.cell(row=r, column=3, value=unit).border = BORDER
        ws.cell(row=r, column=4, value=note).border = BORDER

    # named cell references for the FS formula: B6..B11
    fs_row = start_row + len(rows) + 2
    ws.cell(row=fs_row - 1, column=1, value="Legend: blue text / yellow fill = editable input cell").font = Font(
        name=FONT_NAME, italic=True, size=9)

    ws.cell(row=fs_row, column=1, value="Factor of Safety (baseline, live formula)").font = BOLD
    fs_formula = (
        "=(B6+(B8*B9*COS(RADIANS(B10))^2-B11*B8*B9)*TAN(RADIANS(B7)))"
        "/(B8*B9*SIN(RADIANS(B10))*COS(RADIANS(B10)))"
    )
    fs_cell = ws.cell(row=fs_row, column=2, value=fs_formula)
    fs_cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    fs_cell.number_format = "0.000"
    fs_cell.border = BORDER

    ws.cell(row=fs_row + 1, column=1, value="Interpretation:").font = Font(name=FONT_NAME, italic=True, size=9)
    ws.cell(row=fs_row + 1, column=2,
            value='=IF(B'+str(fs_row)+'<1,"UNSTABLE","STABLE")').font = Font(name=FONT_NAME, italic=True, size=9)

    ws.cell(row=fs_row + 3, column=1,
            value="Design guidance: typical minimum FS targets are 1.3 (static, long-term) and 1.0-1.1 (pseudo-static/seismic).").font = \
        Font(name=FONT_NAME, italic=True, size=9)
    ws.merge_cells(start_row=fs_row + 3, start_column=1, end_row=fs_row + 3, end_column=6)

    autosize(ws, {"A": 30, "B": 14, "C": 10, "D": 45})
    ws.sheet_view.showGridLines = False
    return ws


# ----------------------------------------------------------------------------
# Sheet 2: Sensitivity
# ----------------------------------------------------------------------------
def build_sensitivity_sheet(wb):
    ws = wb.create_sheet("Sensitivity")
    ws["A1"] = "One-at-a-Time (OAT) Sensitivity Analysis"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Each parameter is swept across its realistic range; all others are held at the baseline (Inputs sheet)."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10)

    col_cursor = 1
    input_cell_map = {  # param_name -> Inputs!$B$row
        "cohesion_kPa": "$B$6",
        "friction_angle_deg": "$B$7",
        "unit_weight_kNm3": "$B$8",
        "depth_m": "$B$9",
        "slope_angle_deg": "$B$10",
        "ru": "$B$11",
    }

    table_start_row = 4
    for name, (lo, hi) in PARAM_RANGES.items():
        c0 = col_cursor
        header_col_letter = get_column_letter(c0)
        ws.cell(row=table_start_row - 1, column=c0, value=PARAM_LABELS[name]).font = BOLD
        ws.cell(row=table_start_row, column=c0, value="Value").font = BOLD
        ws.cell(row=table_start_row, column=c0 + 1, value="FS").font = BOLD
        style_header_row(ws, table_start_row, c0, c0 + 1)

        values = np.linspace(lo, hi, N_STEPS)
        for i, v in enumerate(values):
            r = table_start_row + 1 + i
            ws.cell(row=r, column=c0, value=round(float(v), 4))

            refs = dict(input_cell_map)
            refs.pop(name)
            this_cell = f"{header_col_letter}{r}"
            b_c = this_cell if name == "cohesion_kPa" else f"Inputs & Baseline!{refs.get('cohesion_kPa')}" if 'cohesion_kPa' in refs else this_cell
            # Build formula referencing swept column for this param, Inputs sheet for the rest.
            cell_refs = {}
            for key, addr in input_cell_map.items():
                cell_refs[key] = this_cell if key == name else f"'Inputs & Baseline'!{addr}"

            formula = (
                f"=({cell_refs['cohesion_kPa']}+({cell_refs['unit_weight_kNm3']}*{cell_refs['depth_m']}*"
                f"COS(RADIANS({cell_refs['slope_angle_deg']}))^2-{cell_refs['ru']}*{cell_refs['unit_weight_kNm3']}*"
                f"{cell_refs['depth_m']})*TAN(RADIANS({cell_refs['friction_angle_deg']})))"
                f"/({cell_refs['unit_weight_kNm3']}*{cell_refs['depth_m']}*SIN(RADIANS({cell_refs['slope_angle_deg']}))*"
                f"COS(RADIANS({cell_refs['slope_angle_deg']})))"
            )
            fs_cell = ws.cell(row=r, column=c0 + 1, value=formula)
            fs_cell.number_format = "0.000"

        ws.column_dimensions[get_column_letter(c0)].width = 12
        ws.column_dimensions[get_column_letter(c0 + 1)].width = 10
        col_cursor += 3  # one blank column between tables

    # Tornado summary table
    tornado_row = table_start_row + N_STEPS + 4
    ws.cell(row=tornado_row, column=1, value="Tornado Summary (parameter influence, sorted)").font = BOLD
    t_headers = ["Parameter", "Low Value", "High Value", "FS at Low", "FS at High", "Baseline FS", "Swing (|ΔFS|)"]
    for i, h in enumerate(t_headers, start=1):
        ws.cell(row=tornado_row + 1, column=i, value=h)
    style_header_row(ws, tornado_row + 1, 1, len(t_headers))

    tdf = tornado_data().sort_values("swing", ascending=False)
    for i, (_, row) in enumerate(tdf.iterrows()):
        r = tornado_row + 2 + i
        ws.cell(row=r, column=1, value=row["parameter"])
        ws.cell(row=r, column=2, value=round(row["low_value"], 3))
        ws.cell(row=r, column=3, value=round(row["high_value"], 3))
        ws.cell(row=r, column=4, value=round(row["FS_at_low"], 3))
        ws.cell(row=r, column=5, value=round(row["FS_at_high"], 3))
        ws.cell(row=r, column=6, value=round(row["baseline_FS"], 3))
        ws.cell(row=r, column=7, value=round(row["swing"], 3))

    # Charts
    img_row = tornado_row + 2 + len(tdf) + 3
    ws.cell(row=img_row, column=1, value="Charts").font = BOLD
    img1 = XLImage(f"{OUT_DIR}/sensitivity_lines.png")
    img1.width, img1.height = 780, 420
    ws.add_image(img1, f"A{img_row + 1}")

    img2 = XLImage(f"{OUT_DIR}/tornado.png")
    img2.width, img2.height = 720, 430
    ws.add_image(img2, f"A{img_row + 23}")

    ws.sheet_view.showGridLines = False
    return ws


# ----------------------------------------------------------------------------
# Sheet 3: Monte Carlo
# ----------------------------------------------------------------------------
def build_monte_carlo_sheet(wb, df_mc_full, stats_full):
    ws = wb.create_sheet("Monte Carlo")
    ws["A1"] = "Monte Carlo Simulation of Factor of Safety"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Input Distribution Assumptions"
    ws["A3"].font = BOLD
    dist_headers = ["Parameter", "Distribution", "Mean", "Std Dev", "Lower Bound", "Upper Bound"]
    for i, h in enumerate(dist_headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, len(dist_headers))

    dists = default_distributions()
    for i, (name, d) in enumerate(dists.items()):
        r = 5 + i
        ws.cell(row=r, column=1, value=PARAM_LABELS[name])
        ws.cell(row=r, column=2, value="Truncated Normal")
        ws.cell(row=r, column=3, value=d["mean"])
        ws.cell(row=r, column=4, value=d["std"])
        ws.cell(row=r, column=5, value=d["bounds"][0] if d["bounds"][0] is not None else "")
        ws.cell(row=r, column=6, value=d["bounds"][1] if d["bounds"][1] is not None else "")

    # Full-resolution simulation summary (Python-computed; too large to replicate live in-sheet)
    sr = 12
    ws.cell(row=sr, column=1,
            value=f"Full-Resolution Simulation Summary (n = {stats_full['n_trials']:,}, computed in Python "
                  f"via monte_carlo.py — reproducible with the fixed random seed noted in that script)").font = BOLD
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)

    summary_rows = [
        ("Mean FS", stats_full["mean_FS"]),
        ("Std Dev FS", stats_full["std_FS"]),
        ("Min FS", stats_full["min_FS"]),
        ("Max FS", stats_full["max_FS"]),
        ("Probability of Failure, Pf = P(FS<1)", stats_full["probability_of_failure"]),
        ("Reliability Index, beta = (mean(FS)-1)/std(FS)", stats_full["reliability_index_beta"]),
    ]
    for i, (label, val) in enumerate(summary_rows):
        r = sr + 1 + i
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=round(val, 4))
        if "Probability" in label:
            c.number_format = "0.0%"

    ws.cell(row=sr + 8, column=1, value="Percentiles of FS").font = BOLD
    pct_headers = ["P1", "P5", "P10", "P25", "P50", "P75", "P90", "P95", "P99"]
    for i, h in enumerate(pct_headers, start=1):
        ws.cell(row=sr + 9, column=i, value=h)
    style_header_row(ws, sr + 9, 1, len(pct_headers))
    for i, p in enumerate([1, 5, 10, 25, 50, 75, 90, 95, 99]):
        ws.cell(row=sr + 10, column=i + 1, value=round(stats_full["percentiles"][p], 3))

    # Live 2,000-row sample with formula-computed FS (demonstrates the mechanics transparently)
    sample_row = sr + 13
    ws.cell(row=sample_row, column=1,
            value="Worksheet-Live Sample (n = 2,000 trials; FS computed by in-cell formula from the sampled inputs — "
                  "illustrative; full headline results above use n=100,000 computed in Python for stability)").font = BOLD
    ws.merge_cells(start_row=sample_row, start_column=1, end_row=sample_row, end_column=7)

    sample_headers = ["c' (kPa)", "phi' (deg)", "gamma (kN/m3)", "z (m)", "beta (deg)", "r_u", "FS (formula)"]
    for i, h in enumerate(sample_headers, start=1):
        ws.cell(row=sample_row + 1, column=i, value=h)
    style_header_row(ws, sample_row + 1, 1, len(sample_headers))

    n_sample = 2000
    df_sample = df_mc_full.iloc[:n_sample]
    data_start = sample_row + 2
    for i, (_, row) in enumerate(df_sample.iterrows()):
        r = data_start + i
        ws.cell(row=r, column=1, value=round(float(row["cohesion_kPa"]), 3))
        ws.cell(row=r, column=2, value=round(float(row["friction_angle_deg"]), 3))
        ws.cell(row=r, column=3, value=round(float(row["unit_weight_kNm3"]), 3))
        ws.cell(row=r, column=4, value=round(float(row["depth_m"]), 3))
        ws.cell(row=r, column=5, value=round(float(row["slope_angle_deg"]), 3))
        ws.cell(row=r, column=6, value=round(float(row["ru"]), 4))
        formula = (
            f"=(A{r}+(C{r}*D{r}*COS(RADIANS(E{r}))^2-F{r}*C{r}*D{r})*TAN(RADIANS(B{r})))"
            f"/(C{r}*D{r}*SIN(RADIANS(E{r}))*COS(RADIANS(E{r})))"
        )
        ws.cell(row=r, column=7, value=formula).number_format = "0.000"

    data_end = data_start + n_sample - 1
    stat_row = data_end + 2
    ws.cell(row=stat_row, column=1, value="Sample stats (live, from the 2,000-row table above):").font = BOLD
    ws.cell(row=stat_row + 1, column=1, value="Mean FS")
    ws.cell(row=stat_row + 1, column=2, value=f"=AVERAGE(G{data_start}:G{data_end})").number_format = "0.000"
    ws.cell(row=stat_row + 2, column=1, value="Std Dev FS")
    ws.cell(row=stat_row + 2, column=2, value=f"=STDEV(G{data_start}:G{data_end})").number_format = "0.000"
    ws.cell(row=stat_row + 3, column=1, value="Probability of Failure (sample)")
    ws.cell(row=stat_row + 3, column=2,
            value=f"=COUNTIF(G{data_start}:G{data_end},\"<1\")/COUNT(G{data_start}:G{data_end})").number_format = "0.0%"

    # Charts placed near the top-right, away from the long data table
    img1 = XLImage(f"{OUT_DIR}/mc_histogram.png")
    img1.width, img1.height = 680, 420
    ws.add_image(img1, "I3")

    img2 = XLImage(f"{OUT_DIR}/mc_convergence.png")
    img2.width, img2.height = 680, 340
    ws.add_image(img2, "I25")

    autosize(ws, {"A": 34, "B": 12, "C": 12, "D": 12, "E": 12, "F": 10, "G": 12})
    ws.sheet_view.showGridLines = False
    return ws


# ----------------------------------------------------------------------------
# Sheet 4: Risk Optimization
# ----------------------------------------------------------------------------
def build_optimization_sheet(wb, df_sweep):
    ws = wb.create_sheet("Risk Optimization")
    ws["A1"] = "Risk-Based Optimization of Dump Slope Angle"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Full Monte Carlo simulation (n = 20,000 per angle) re-run at each candidate slope angle, "
                "holding it fixed while all other geotechnical inputs retain their baseline distributions.")
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10)
    ws.merge_cells("A2:F2")

    headers = ["Slope Angle (deg)", "Mean FS", "Std Dev FS", "Probability of Failure", "Reliability Index"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, len(headers))

    for i, (_, row) in enumerate(df_sweep.iterrows()):
        r = 5 + i
        ws.cell(row=r, column=1, value=row["slope_angle_deg"])
        ws.cell(row=r, column=2, value=round(row["mean_FS"], 3))
        ws.cell(row=r, column=3, value=round(row["std_FS"], 3))
        pc = ws.cell(row=r, column=4, value=round(row["probability_of_failure"], 5))
        pc.number_format = "0.0%"
        ws.cell(row=r, column=5, value=round(row["reliability_index_beta"], 3))

    rec_row = 5 + len(df_sweep) + 2
    ws.cell(row=rec_row, column=1, value="Recommended Slope Angles by Risk Target").font = BOLD
    rec_headers = ["Target Pf", "Recommended Slope Angle (deg)", "Achieved Pf"]
    for i, h in enumerate(rec_headers, start=1):
        ws.cell(row=rec_row + 1, column=i, value=h)
    style_header_row(ws, rec_row + 1, 1, len(rec_headers))

    rec_df = recommended_angles(df_sweep, TARGET_PF)
    for i, (_, row) in enumerate(rec_df.iterrows()):
        r = rec_row + 2 + i
        ws.cell(row=r, column=1, value=row["target_Pf"]).number_format = "0%"
        ws.cell(row=r, column=2, value=row["recommended_slope_angle_deg"])
        c = ws.cell(row=r, column=3, value=row["achieved_Pf"])
        c.number_format = "0.0%"

    img = XLImage(f"{OUT_DIR}/risk_optimization.png")
    img.width, img.height = 760, 440
    ws.add_image(img, f"A{rec_row + 2 + len(rec_df) + 3}")

    autosize(ws, {"A": 22, "B": 14, "C": 14, "D": 20, "E": 18})
    ws.sheet_view.showGridLines = False
    return ws


# ----------------------------------------------------------------------------
# Sheet 5: Notes & Methodology
# ----------------------------------------------------------------------------
def build_notes_sheet(wb):
    ws = wb.create_sheet("Notes & Methodology")
    ws["A1"] = "Notes & Methodology"
    ws["A1"].font = TITLE_FONT

    notes = [
        "",
        "1. MODEL",
        "Infinite-slope method with seepage parallel to the slope face, using the pore-pressure ratio r_u "
        "(Bishop & Morgenstern, 1960). This is a standard first-pass screening tool for planar / near-surface "
        "failures in engineered waste rock dumps, distinct from deeper rotational failures which would require "
        "a circular slip-surface (e.g., Bishop's Simplified) slice analysis.",
        "",
        "FS = [c' + (gamma*z*cos^2(beta) - r_u*gamma*z) * tan(phi')] / [gamma*z*sin(beta)*cos(beta)]",
        "",
        "2. BASELINE CASE",
        "The baseline case represents a marginally stable design (FS ~ 1.18) so that both the sensitivity study "
        "and the Monte Carlo simulation show meaningful movement across the FS = 1 failure threshold in either "
        "direction. A more conservative as-built design would target FS >= 1.3 (static).",
        "",
        "3. SENSITIVITY ANALYSIS",
        "One-at-a-time (OAT) sweeps show slope angle and pore pressure ratio r_u are the dominant controls on "
        "FS, followed by friction angle. Cohesion, depth, and unit weight have comparatively minor influence "
        "over their realistic ranges for this dump geometry.",
        "",
        "4. MONTE CARLO SIMULATION",
        "Each input is modeled as an independent, truncated-normal random variable reflecting realistic "
        "field/lab variability. 100,000 trials were run to obtain a converged probability of failure "
        "(Pf = P(FS<1)) and reliability index (beta). Independence between inputs is a simplifying assumption; "
        "in practice, c' and phi' can be correlated (typically negatively) — a refinement noted as a limitation "
        "below.",
        "",
        "5. RISK-BASED OPTIMIZATION",
        "The slope angle sweep re-runs the full Monte Carlo simulation with the slope angle fixed at each "
        "candidate value, isolating the one design lever (mine planners control face angle / benching) from "
        "the geotechnical uncertainty that cannot be engineered away. This produces a Pf-vs-angle design curve, "
        "letting the dump be designed to an explicit risk target (e.g., Pf <= 5%) rather than a single "
        "deterministic FS target.",
        "",
        "6. LIMITATIONS",
        "- Infinite-slope method assumes an infinitely long slope and a failure plane parallel to the face; "
        "it does not capture deep rotational or wedge failures, toe buckling, or foundation failure.",
        "- Inputs are assumed statistically independent; correlation between c' and phi' (commonly negative) "
        "would generally reduce the computed Pf and could be added via a correlated sampling scheme.",
        "- Pore pressure is simplified to a single ratio r_u; a full seepage analysis (e.g., finite element) "
        "would better capture a real phreatic surface and seasonal variation.",
        "- Results are illustrative of method and workflow, not a certified design for an actual site; any "
        "real waste dump design must follow applicable regulatory guidance and be reviewed by a qualified "
        "geotechnical engineer of record.",
        "",
        "7. FILES",
        "core.py - deterministic FS model  |  sensitivity.py - OAT sweeps & tornado data  |  "
        "monte_carlo.py - probabilistic simulation  |  optimization.py - slope-angle risk optimization  |  "
        "make_plots.py - all figures  |  build_excel.py - this workbook",
    ]
    for i, line in enumerate(notes):
        r = 2 + i
        cell = ws.cell(row=r, column=1, value=line)
        if line.strip().startswith(tuple(f"{n}." for n in range(1, 8))):
            cell.font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 15 if len(line) < 90 else 30

    autosize(ws, {"A": 110})
    ws.sheet_view.showGridLines = False
    return ws


def main():
    wb = Workbook()
    build_inputs_sheet(wb)
    build_sensitivity_sheet(wb)

    df_mc_full = run_simulation(100_000)
    stats_full = summarize(df_mc_full)
    build_monte_carlo_sheet(wb, df_mc_full, stats_full)

    df_sweep = sweep_slope_angle()
    build_optimization_sheet(wb, df_sweep)

    build_notes_sheet(wb)

    # Order sheets sensibly
    order = ["Inputs & Baseline", "Sensitivity", "Monte Carlo", "Risk Optimization", "Notes & Methodology"]
    wb._sheets.sort(key=lambda s: order.index(s.title))

    wb.save(WB_PATH)
    print(f"Saved workbook to {WB_PATH}")


if __name__ == "__main__":
    main()
